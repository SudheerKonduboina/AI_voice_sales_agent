"""Generates leads_template.xlsx -- the CRM sheet the voice agent reads from and writes to."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

COLUMNS = [
    "Lead ID",
    "Name",
    "Phone",
    "Email",
    "Status",
    "Qualification",
    "Conversation Summary",
    "Customer Requirements",
    "Objections Raised",
    "Follow-up Date",
    "Meeting Date & Time",
    "Last Contacted",
    "Call Attempts",
    "Notes",
]

STATUS_OPTIONS = '"Pending,Called,Booked,Not Interested,No Answer,Opted Out"'
QUAL_OPTIONS = '"Hot,Warm,Cold,Unqualified,Not Yet Assessed"'

wb = Workbook()
ws = wb.active
ws.title = "Leads"

header_font = Font(name="Arial", bold=True, color="FFFFFF")
header_fill = PatternFill("solid", start_color="2F5496")
for col, name in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=1, column=col, value=name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 30

widths = [10, 16, 15, 22, 14, 16, 32, 28, 28, 14, 18, 18, 12, 24]
for col, w in zip(range(1, len(COLUMNS) + 1), widths, strict=True):
    ws.column_dimensions[chr(64 + col)].width = w

demo_leads = [
    [
        "L001",
        "Rahul Mehta",
        "+91-9876543210",
        "rahul.mehta@example.com",
        "Pending",
        "Not Yet Assessed",
        "",
        "",
        "",
        "",
        "",
        "",
        0,
        "",
    ],
    [
        "L002",
        "Priya Nair",
        "+91-9823456712",
        "priya.nair@example.com",
        "Booked",
        "Hot",
        "Interested in premium plan, wants onboarding support",
        "Team of 12, needs multi-user access",
        "Price a bit high initially",
        "",
        "2026-07-15 11:00",
        "2026-07-08 10:32",
        1,
        "Very engaged, follow up before meeting",
    ],
    [
        "L003",
        "Arjun Desai",
        "+91-9812345678",
        "arjun.desai@example.com",
        "Not Interested",
        "Unqualified",
        "Currently using a competitor, happy with it",
        "N/A",
        "Already has a solution",
        "2026-09-01",
        "",
        "2026-07-07 15:10",
        1,
        "Revisit in 2 months",
    ],
    [
        "L004",
        "Sneha Kulkarni",
        "+91-9845678901",
        "sneha.k@example.com",
        "No Answer",
        "Not Yet Assessed",
        "",
        "",
        "",
        "",
        "",
        "2026-07-08 09:15",
        2,
        "Try again in evening",
    ],
    [
        "L005",
        "Vikram Rao",
        "+91-9834567890",
        "vikram.rao@example.com",
        "Pending",
        "Not Yet Assessed",
        "",
        "",
        "",
        "",
        "",
        "",
        0,
        "",
    ],
]
for r, row in enumerate(demo_leads, start=2):
    for c, val in enumerate(row, start=1):
        ws.cell(row=r, column=c, value=val)

dv_status = DataValidation(type="list", formula1=STATUS_OPTIONS, allow_blank=True)
dv_qual = DataValidation(type="list", formula1=QUAL_OPTIONS, allow_blank=True)
ws.add_data_validation(dv_status)
ws.add_data_validation(dv_qual)
dv_status.add("E2:E1000")
dv_qual.add("F2:F1000")

ws.freeze_panes = "A2"

wb.save("leads_template.xlsx")
print("Saved leads_template.xlsx")
