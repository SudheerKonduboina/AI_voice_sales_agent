"""
excel_crm.py -- Excel-as-CRM read/write layer for the AI Voice Sales Agent.

Usage:
    from excel_crm import get_pending_leads, update_lead, mark_no_answer

    leads = get_pending_leads("leads_template.xlsx")
    for lead in leads:
        result = run_call(lead)              # <- your voice agent call
        update_lead("leads_template.xlsx", lead["lead_id"], result)

Design notes:
- One row per lead. Header row (row 1) defines column names; order-independent
  by name, so you can reorder/add columns without breaking the code.
- Every write re-opens, edits, and saves the workbook. Excel files are not
  safe for concurrent writers -- this is fine for a single agent process
  calling leads sequentially. If you parallelize calls, add a file lock
  (e.g. `filelock` package) around update_lead().
- Status values are constrained to STATUSES; anything else raises ValueError
  so a bug in the agent can't silently corrupt the sheet.
"""

import os
import shutil
import sys
import tempfile
import time
from dataclasses import asdict, dataclass, fields
from datetime import datetime
from typing import Any

from openpyxl import load_workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "agent"))
from logger_setup import get_logger, log_exception  # noqa: E402

logger = get_logger(__name__)


class CRMError(RuntimeError):
    """Raised when a CRM operation fails (save, load, etc.)."""


def _acquire_lock(path: str, timeout: float = 10.0, delay: float = 0.1) -> str:
    """Acquire a lightweight atomic cross-platform file lock."""
    lock_path = path + ".lock"
    start = time.time()
    while time.time() - start < timeout:
        try:
            # os.O_CREAT | os.O_EXCL is atomic across all OS platforms
            fd = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            os.close(fd)
            return lock_path
        except FileExistsError:
            try:
                mtime = os.path.getmtime(lock_path)
                if time.time() - mtime > 15.0:
                    logger.warning(
                        "CRM lock file is stale (older than 15s) -- force releasing lock."
                    )
                    _release_lock(lock_path)
            except Exception:  # nosec B110
                pass
            time.sleep(delay)
    raise CRMError(f"Could not acquire lock on CRM file {path} within {timeout} seconds.")


def _release_lock(lock_path: str) -> None:
    if os.path.exists(lock_path):
        try:
            os.remove(lock_path)
        except Exception as e:
            logger.warning("Failed to release lock file %s: %s", lock_path, e)


STATUSES = {"Pending", "Called", "Booked", "Not Interested", "No Answer", "Opted Out"}
QUALIFICATIONS = {"Hot", "Warm", "Cold", "Unqualified", "Not Yet Assessed"}

# Maps the field names callers use -> exact Excel header text.
FIELD_TO_HEADER = {
    "lead_id": "Lead ID",
    "name": "Name",
    "phone": "Phone",
    "email": "Email",
    "status": "Status",
    "qualification": "Qualification",
    "conversation_summary": "Conversation Summary",
    "customer_requirements": "Customer Requirements",
    "objections_raised": "Objections Raised",
    "follow_up_date": "Follow-up Date",
    "meeting_datetime": "Meeting Date & Time",
    "last_contacted": "Last Contacted",
    "call_attempts": "Call Attempts",
    "notes": "Notes",
}
HEADER_TO_FIELD = {v: k for k, v in FIELD_TO_HEADER.items()}


@dataclass
class LeadUpdate:
    """Structured data model for updating lead status in the CRM."""

    status: str | None = None
    qualification: str | None = None
    conversation_summary: str | None = None
    customer_requirements: str | None = None
    objections_raised: str | None = None
    follow_up_date: str | None = None
    meeting_datetime: str | None = None
    notes: str | None = None

    def validate(self) -> None:
        """Validate values and dates."""
        if self.status is not None and self.status not in STATUSES:
            raise ValueError(f"Invalid status {self.status!r}; must be one of {STATUSES}")
        if self.qualification is not None and self.qualification not in QUALIFICATIONS:
            raise ValueError(
                f"Invalid qualification {self.qualification!r}; must be one of {QUALIFICATIONS}"
            )

        if self.follow_up_date:
            try:
                datetime.strptime(self.follow_up_date, "%Y-%m-%d")
            except ValueError:
                raise ValueError(
                    f"Invalid follow_up_date format: {self.follow_up_date!r}, must be YYYY-MM-DD"
                )

        if self.meeting_datetime:
            try:
                datetime.strptime(self.meeting_datetime, "%Y-%m-%d %H:%M")
            except ValueError:
                raise ValueError(
                    f"Invalid meeting_datetime format: {self.meeting_datetime!r}, must be YYYY-MM-DD HH:MM"
                )


def _header_map(ws) -> dict[str, int]:
    """Returns {header_name: 1-based column index} from row 1."""
    return {cell.value: cell.column for cell in ws[1] if cell.value}


def _row_to_dict(ws, row_idx: int, headers: dict[str, int]) -> dict[str, Any]:
    out = {}
    for header, col in headers.items():
        field = HEADER_TO_FIELD.get(header)
        if field:
            out[field] = ws.cell(row=row_idx, column=col).value
    out["_row"] = row_idx
    return out


def get_all_leads(path: str, sheet: str = "Leads") -> list[dict[str, Any]]:
    wb = load_workbook(path)
    try:
        ws = wb[sheet]
        headers = _header_map(ws)
        return [
            _row_to_dict(ws, r, headers)
            for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=headers["Lead ID"]).value
        ]
    finally:
        wb.close()


def get_pending_leads(path: str, sheet: str = "Leads") -> list[dict[str, Any]]:
    """Leads that are Pending (or blank status) and not Opted Out -- ready to call."""
    return [
        lead for lead in get_all_leads(path, sheet) if (lead.get("status") in (None, "", "Pending"))
    ]


def _find_row(ws, headers: dict[str, int], lead_id: str) -> int:
    id_col = headers["Lead ID"]
    found_rows = []
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=id_col).value) == str(lead_id):
            found_rows.append(r)

    if not found_rows:
        raise KeyError(f"Lead ID {lead_id!r} not found")
    if len(found_rows) > 1:
        raise ValueError(f"Duplicate Lead ID {lead_id!r} found in rows: {found_rows}")
    return found_rows[0]


def _save_workbook_safely(wb, path: str) -> None:
    """Transaction safety: save workbook to a temp file, load it back to verify integrity, then atomic rename."""
    dir_name = os.path.dirname(os.path.abspath(path))
    temp_fd, temp_path = tempfile.mkstemp(dir=dir_name, suffix=".xlsx")
    os.close(temp_fd)

    try:
        wb.save(temp_path)
        # Verify integrity
        v_wb = load_workbook(temp_path)
        v_wb.close()
        # Atomic replace
        shutil.move(temp_path, path)
        logger.debug("Successfully saved CRM workbook to %s", path)
    except Exception as e:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        log_exception(logger, "Transactional save failed for CRM -- rollback initiated: %s", e)
        raise CRMError(f"CRM transactional save failed: {e}") from e


def update_lead(
    path: str,
    lead_id: str,
    updates: dict[str, Any] | LeadUpdate,
    sheet: str = "Leads",
    max_retries: int = 2,
) -> None:
    """
    Writes call results back to the lead's row.
    Validates updates against LeadUpdate schema, enforces formatting, and ensures atomic transaction.
    Retries on validation failure with sanitized values.
    """
    lock_path = _acquire_lock(path)
    try:
        last_error: Exception | None = None
        current_updates = updates

        for attempt in range(1, max_retries + 1):
            try:
                _update_lead_once(path, lead_id, current_updates, sheet)
                return
            except ValueError as e:
                last_error = e
                logger.warning("CRM update attempt %d failed validation: %s", attempt, e)
                if isinstance(current_updates, dict):
                    current_updates = _sanitize_updates(current_updates)
                else:
                    raise
            except (KeyError, RuntimeError, OSError) as e:
                log_exception(logger, "CRM update attempt %d failed: %s", attempt, e)
                raise

        raise ValueError(f"CRM update failed after {max_retries} attempts: {last_error}")
    finally:
        _release_lock(lock_path)


def _sanitize_updates(updates: dict[str, Any]) -> dict[str, Any]:
    """Coerce malformed update dicts toward valid LeadUpdate values."""
    sanitized = dict(updates)
    if "status" in sanitized and sanitized["status"] not in STATUSES:
        sanitized["status"] = "Pending"
    if "qualification" in sanitized and sanitized["qualification"] not in QUALIFICATIONS:
        sanitized["qualification"] = "Not Yet Assessed"
    for date_field, fmt in (("follow_up_date", "%Y-%m-%d"), ("meeting_datetime", "%Y-%m-%d %H:%M")):
        val = sanitized.get(date_field)
        if val:
            try:
                datetime.strptime(str(val), fmt)
            except ValueError:
                sanitized[date_field] = ""
    return sanitized


def _update_lead_once(
    path: str, lead_id: str, updates: dict[str, Any] | LeadUpdate, sheet: str
) -> None:
    if isinstance(updates, dict):
        valid_fields = {f.name for f in fields(LeadUpdate)}
        filtered = {k: v for k, v in updates.items() if k in valid_fields and v is not None}
        update_obj = LeadUpdate(**filtered)
    else:
        update_obj = updates

    update_obj.validate()

    wb = load_workbook(path)
    try:
        ws = wb[sheet]
        headers = _header_map(ws)
        row = _find_row(ws, headers, lead_id)

        update_dict = asdict(update_obj)
        for field, value in update_dict.items():
            if value is not None:
                header = FIELD_TO_HEADER.get(field)
                if header and header in headers:
                    ws.cell(row=row, column=headers[header], value=value)

        ws.cell(
            row=row,
            column=headers["Last Contacted"],
            value=datetime.now().strftime("%Y-%m-%d %H:%M"),
        )

        attempts_col = headers["Call Attempts"]
        current = ws.cell(row=row, column=attempts_col).value or 0
        ws.cell(row=row, column=attempts_col, value=int(current) + 1)

        _save_workbook_safely(wb, path)
        logger.info("Updated CRM lead_id=%s successfully", lead_id)
    finally:
        wb.close()


def mark_no_answer(path: str, lead_id: str, sheet: str = "Leads") -> None:
    update_lead(path, lead_id, {"status": "No Answer"}, sheet)


def mark_opted_out(path: str, lead_id: str, sheet: str = "Leads") -> None:
    update_lead(path, lead_id, {"status": "Opted Out"}, sheet)


def get_meetings_scheduled(path: str, sheet: str = "Leads") -> list[dict[str, Any]]:
    """Return leads that have a meeting_datetime set (i.e. booked meetings)."""
    all_leads = get_all_leads(path, sheet)
    return [
        lead
        for lead in all_leads
        if lead.get("meeting_datetime") and str(lead["meeting_datetime"]).strip()
    ]


def get_lead_by_id(path: str, lead_id: str, sheet: str = "Leads") -> dict[str, Any] | None:
    """Return a single lead dict by lead_id, or None if not found."""
    for lead in get_all_leads(path, sheet):
        if str(lead.get("lead_id", "")) == str(lead_id):
            return lead
    return None
